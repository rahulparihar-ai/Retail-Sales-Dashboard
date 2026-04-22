"""Master Excel Dashboard — uses retail_sales_final.csv (300 rows, 2020-2022)"""
import csv, calendar
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def bd():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=10, color="222222"): return Font(bold=bold, size=sz, color=color, name="Calibri")
def wfnt(sz=10): return Font(bold=True, size=sz, color="FFFFFF", name="Calibri")
def cs(c, val, bold=False, sz=10, color="222222", bg=None, align="left", num_fmt=None):
    c.value = val; c.font = fnt(bold, sz, color); c.border = bd()
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg: c.fill = fill(bg)
    if num_fmt: c.number_format = num_fmt
def hdr(ws, row, col, text, bg="1F4E79", sz=10):
    c = ws.cell(row=row, column=col)
    cs(c, text, bold=True, sz=sz, color="FFFFFF", bg=bg, align="center")

# ── Load data ─────────────────────────────────────────────────────────────────
rows = []
with open(r"e:\BEE\Section1_Dataset\retail_sales_final.csv", newline='', encoding='utf-8') as f:
    for r in csv.DictReader(f): rows.append(r)
for r in rows:
    r['Sales'] = float(r['Sales']); r['Profit'] = float(r['Profit'])
    r['Quantity'] = int(r['Quantity']); r['Discount'] = float(r['Discount'])
    p = r['Order Date'].split('-')
    r['Day'], r['Month'], r['Year'] = int(p[0]), int(p[1]), int(p[2])
    r['MonthName'] = calendar.month_abbr[r['Month']]

CATS = sorted(set(r['Category'] for r in rows))
REGIONS = ['East', 'West', 'Central', 'South']
YEARS = sorted(set(r['Year'] for r in rows))
MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

cat_reg_s = defaultdict(lambda: defaultdict(float))
cat_reg_p = defaultdict(lambda: defaultdict(float))
yr_mon_s  = defaultdict(lambda: defaultdict(float))
prod_p    = defaultdict(float)
seg_s     = defaultdict(float)
yr_s      = defaultdict(float)
yr_p      = defaultdict(float)
subcat_p  = defaultdict(float)
for r in rows:
    cat_reg_s[r['Category']][r['Region']] += r['Sales']
    cat_reg_p[r['Category']][r['Region']] += r['Profit']
    yr_mon_s[r['Year']][r['MonthName']]   += r['Sales']
    prod_p[r['Product Name']]             += r['Profit']
    seg_s[r['Segment']]                   += r['Sales']
    yr_s[r['Year']]  += r['Sales']
    yr_p[r['Year']]  += r['Profit']
    subcat_p[r['Sub-Category']] += r['Profit']

top10   = sorted(prod_p.items(),  key=lambda x: x[1], reverse=True)[:10]
total_s = sum(r['Sales']  for r in rows)
total_p = sum(r['Profit'] for r in rows)
total_o = len(rows)
total_q = sum(r['Quantity'] for r in rows)
margin  = total_p / total_s
s2020, s2021, s2022 = yr_s[2020], yr_s[2021], yr_s[2022]
yoy2021 = (s2021 - s2020) / s2020 * 100
yoy2022 = (s2022 - s2021) / s2021 * 100

wb = Workbook()

# ═══ SHEET: SalesData ═════════════════════════════════════════════════════════
ws_d = wb.active; ws_d.title = "SalesData"
ws_d.sheet_properties.tabColor = "2E75B6"
COLS = ["Order ID","Order Date","Ship Date","Customer Name","Segment","Region",
        "State","Category","Sub-Category","Product Name","Sales","Quantity","Discount","Profit"]
WIDTHS = [17,13,13,18,13,10,15,17,14,34,11,9,10,11]
for ci,(h,w) in enumerate(zip(COLS,WIDTHS),1):
    c = ws_d.cell(row=1, column=ci, value=h)
    c.fill = fill("1F4E79"); c.font = wfnt(10)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd()
    ws_d.column_dimensions[get_column_letter(ci)].width = w
ws_d.row_dimensions[1].height = 28
for ri, r in enumerate(rows, 2):
    bg = "EBF3FB" if ri%2==0 else "FFFFFF"
    vals = [r['Order ID'],r['Order Date'],r['Ship Date'],r['Customer Name'],r['Segment'],
            r['Region'],r['State'],r['Category'],r['Sub-Category'],r['Product Name'],
            r['Sales'],r['Quantity'],r['Discount'],r['Profit']]
    for ci, v in enumerate(vals, 1):
        c = ws_d.cell(row=ri, column=ci, value=v)
        c.font = fnt(sz=9); c.border = bd(); c.fill = fill(bg)
        c.alignment = Alignment(vertical="center")
        if ci in (11,14): c.number_format='"$"#,##0.00'; c.alignment=Alignment(horizontal="right",vertical="center")
        elif ci==13: c.number_format='0%'
tbl = Table(displayName="SalesData", ref=f"A1:{get_column_letter(14)}{len(rows)+1}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws_d.add_table(tbl); ws_d.freeze_panes = "A2"

# ═══ SHEET: PT_Category ═══════════════════════════════════════════════════════
ws_c = wb.create_sheet("PT_Category"); ws_c.sheet_properties.tabColor = "375623"
ws_c.merge_cells("A1:H1"); c = ws_c["A1"]
c.value = "Sales & Profit by Category and Region (2020-2022)"; c.fill = fill("1F4E79")
c.font = wfnt(13); c.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[1].height = 30
for ci, lbl in enumerate(["Category","Metric"] + REGIONS + ["Grand Total"], 1):
    hdr(ws_c, 2, ci, lbl)
row = 3
for cat in CATS:
    for metric, data in [("Sales $", cat_reg_s), ("Profit $", cat_reg_p)]:
        bg = "EBF3FB" if row%2==0 else "FFFFFF"
        cs(ws_c.cell(row=row,column=1), cat, bold=True, bg=bg)
        cs(ws_c.cell(row=row,column=2), metric, bg=bg)
        total = 0
        for ci2, reg in enumerate(REGIONS, 3):
            v = data[cat][reg]; total += v
            c = ws_c.cell(row=row, column=ci2)
            cs(c, v, bg=bg, align="right", num_fmt='"$"#,##0.00')
        c = ws_c.cell(row=row, column=8)
        cs(c, total, bold=True, bg=bg, align="right", num_fmt='"$"#,##0.00')
        row += 1
for ci2, w in enumerate([18,10,12,12,12,12,14],1):
    ws_c.column_dimensions[get_column_letter(ci2)].width = w
# Chart
ch1 = BarChart(); ch1.type = "col"; ch1.title = "Sales by Category & Region"
ch1.y_axis.title = "Sales ($)"; ch1.style = 10; ch1.width = 24; ch1.height = 14
for idx, cat in enumerate(CATS):
    dr = Reference(ws_c, min_col=3, max_col=6, min_row=3+idx*2, max_row=3+idx*2)
    ch1.add_data(dr); ch1.series[-1].title = SeriesLabel(v=cat)
ch1.set_categories(Reference(ws_c, min_col=3, max_col=6, min_row=2))
ws_c.add_chart(ch1, "A15")

# ═══ SHEET: PT_YearlyComparison ═══════════════════════════════════════════════
ws_y = wb.create_sheet("PT_YearlyComparison"); ws_y.sheet_properties.tabColor = "7030A0"
ws_y.merge_cells("A1:D1"); c = ws_y["A1"]
c.value = "Year-over-Year Sales Comparison (2020 / 2021 / 2022)"; c.fill = fill("1F4E79")
c.font = wfnt(13); c.alignment = Alignment(horizontal="center", vertical="center")
ws_y.row_dimensions[1].height = 30
for ci2, lbl in enumerate(["Month","2020","2021","2022"], 1): hdr(ws_y, 2, ci2, lbl)
for ri2, m in enumerate(MONTHS, 3):
    bg = "EBF3FB" if ri2%2==0 else "FFFFFF"
    cs(ws_y.cell(row=ri2,column=1), m, bg=bg, align="center")
    for ci2, yr in enumerate([2020,2021,2022], 2):
        v = yr_mon_s[yr].get(m, 0)
        cs(ws_y.cell(row=ri2,column=ci2), v, bg=bg, align="right", num_fmt='"$"#,##0.00')
r_tot = 15
cs(ws_y.cell(row=r_tot,column=1),"TOTAL",bold=True,bg="1F4E79",color="FFFFFF",align="center")
for ci2, (yr, s) in enumerate([(2020,s2020),(2021,s2021),(2022,s2022)], 2):
    cs(ws_y.cell(row=r_tot,column=ci2), s, bold=True, bg="1F4E79", color="FFFFFF", align="right", num_fmt='"$"#,##0.00')
r_yoy = 16
cs(ws_y.cell(row=r_yoy,column=1),"YoY Growth",bold=True,bg="375623",color="FFFFFF",align="center")
cs(ws_y.cell(row=r_yoy,column=2),"—",bold=True,bg="375623",color="FFFFFF",align="center")
cs(ws_y.cell(row=r_yoy,column=3),yoy2021/100,bold=True,bg="375623",color="FFFFFF",align="right",num_fmt='+0.0%;-0.0%')
cs(ws_y.cell(row=r_yoy,column=4),yoy2022/100,bold=True,bg="375623",color="FFFFFF",align="right",num_fmt='+0.0%;-0.0%')
for ci2, w in enumerate([12,15,15,15],1): ws_y.column_dimensions[get_column_letter(ci2)].width = w
# YoY line chart
ch2 = LineChart(); ch2.title = "Monthly Sales: 2020 vs 2021 vs 2022"
ch2.y_axis.title = "Sales ($)"; ch2.style = 10; ch2.width = 24; ch2.height = 14
colors = ["4472C4","ED7D31","A9D18E"]
for ci2, (yr, clr) in enumerate(zip([2020,2021,2022], colors), 2):
    ch2.add_data(Reference(ws_y,min_col=ci2,max_col=ci2,min_row=2,max_row=14),titles_from_data=True)
    ch2.series[-1].graphicalProperties.line.solidFill = clr
    ch2.series[-1].graphicalProperties.line.width = 25000
ch2.set_categories(Reference(ws_y,min_col=1,max_col=1,min_row=3,max_row=14))
ws_y.add_chart(ch2, "F2")

# ═══ SHEET: PT_TopProducts ════════════════════════════════════════════════════
ws_t = wb.create_sheet("PT_TopProducts"); ws_t.sheet_properties.tabColor = "833C00"
ws_t.merge_cells("A1:C1"); c = ws_t["A1"]
c.value = "Top 10 Products by Profit (2020-2022)"; c.fill = fill("1F4E79")
c.font = wfnt(13); c.alignment = Alignment(horizontal="center", vertical="center")
ws_t.row_dimensions[1].height = 30
for ci2, lbl in enumerate(["Rank","Product Name","Profit ($)"],1): hdr(ws_t, 2, ci2, lbl)
for ri2,(prod,profit) in enumerate(top10, 3):
    bg = "EBF3FB" if ri2%2==0 else "FFFFFF"
    cs(ws_t.cell(row=ri2,column=1), ri2-2, bg=bg, align="center")
    cs(ws_t.cell(row=ri2,column=2), prod[:42]+"…" if len(prod)>42 else prod, bg=bg)
    cs(ws_t.cell(row=ri2,column=3), profit, bg=bg, align="right", num_fmt='"$"#,##0.00')
ws_t.column_dimensions["A"].width=7; ws_t.column_dimensions["B"].width=44; ws_t.column_dimensions["C"].width=14
ch3 = BarChart(); ch3.type = "bar"; ch3.title = "Top 10 Products by Profit"
ch3.style = 10; ch3.width = 26; ch3.height = 16
ch3.add_data(Reference(ws_t,min_col=3,max_col=3,min_row=2,max_row=12),titles_from_data=True)
ch3.set_categories(Reference(ws_t,min_col=2,max_col=2,min_row=3,max_row=12))
ws_t.add_chart(ch3, "E2")

# ═══ SHEET: PT_SubCategory ════════════════════════════════════════════════════
ws_sc = wb.create_sheet("PT_SubCategory"); ws_sc.sheet_properties.tabColor = "C55A11"
ws_sc.merge_cells("A1:B1"); c = ws_sc["A1"]
c.value = "Profit by Sub-Category (All Years)"; c.fill = fill("1F4E79")
c.font = wfnt(13); c.alignment = Alignment(horizontal="center",vertical="center")
ws_sc.row_dimensions[1].height = 30
hdr(ws_sc,2,1,"Sub-Category"); hdr(ws_sc,2,2,"Total Profit ($)")
sorted_sc = sorted(subcat_p.items(), key=lambda x: x[1], reverse=True)
for ri2,(sc,prof) in enumerate(sorted_sc, 3):
    bg = "EBF3FB" if ri2%2==0 else "FFFFFF"
    fg = "CC0000" if prof < 0 else "222222"
    cs(ws_sc.cell(row=ri2,column=1), sc, bg=bg)
    c = ws_sc.cell(row=ri2, column=2); cs(c, prof, bg=bg, align="right", num_fmt='"$"#,##0.00')
    c.font = fnt(sz=10, color=fg, bold=(prof < 0))
ws_sc.column_dimensions["A"].width = 20; ws_sc.column_dimensions["B"].width = 18

# ═══ SHEET: DASHBOARD ═════════════════════════════════════════════════════════
ws_db = wb.create_sheet("DASHBOARD", 0)
ws_db.sheet_properties.tabColor = "1F4E79"
ws_db.sheet_view.showGridLines = False
for i in range(1,38): ws_db.column_dimensions[get_column_letter(i)].width = 4.0
for ri2 in range(1,75): ws_db.row_dimensions[ri2].height = 17

def dm(r1,c1,r2,c2,val,bg="1F4E79",fsz=11,bold=True,align="center",color="FFFFFF"):
    ws_db.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    c = ws_db.cell(row=r1,column=c1,value=val)
    c.fill = fill(bg); c.font = Font(bold=bold,size=fsz,color=color,name="Calibri")
    c.alignment = Alignment(horizontal=align,vertical="center",wrap_text=True)

# Header
dm(1,2,3,35,"STORE SALES PERFORMANCE DASHBOARD   |   Group 4 — BEE Lab   |   2020–2022","1F4E79",17)
dm(4,2,4,35,"Retail Sales Analytics   |   300 Orders   |   3 Years of Data   |   All Regions & Segments","2E75B6",10,False)

# KPI Cards row 1
kpis1 = [("Total Sales","${:,.0f}".format(total_s),"1F4E79"),
          ("Total Profit","${:,.0f}".format(total_p),"375623"),
          ("Profit Margin","{:.1f}%".format(margin*100),"7030A0"),
          ("Total Orders",str(total_o),"C55A11"),
          ("Total Units",str(total_q),"833C00")]
starts1 = [2,9,16,23,30]
for (lbl,val,hx),sc in zip(kpis1,starts1):
    dm(6,sc,6,sc+6,lbl,hx,9); dm(7,sc,9,sc+6,val,hx,16)

# KPI Cards row 2 — YoY
kpis2=[("2020 Sales","${:,.0f}".format(s2020),"2E75B6"),
       ("2021 Sales","${:,.0f}".format(s2021),"2E75B6"),
       ("2021 YoY Growth","+{:.1f}%".format(yoy2021),"375623"),
       ("2022 Sales","${:,.0f}".format(s2022),"2E75B6"),
       ("2022 YoY Growth","{:+.1f}%".format(yoy2022),"7030A0")]
for (lbl,val,hx),sc in zip(kpis2,starts1):
    dm(11,sc,11,sc+6,lbl,hx,9); dm(12,sc,14,sc+6,val,hx,14)

ws_db.row_dimensions[6].height=17; ws_db.row_dimensions[7].height=22
ws_db.row_dimensions[8].height=22; ws_db.row_dimensions[9].height=22
ws_db.row_dimensions[11].height=17; ws_db.row_dimensions[12].height=20
ws_db.row_dimensions[13].height=20; ws_db.row_dimensions[14].height=20

dm(16,2,16,35,"Category & Regional Analysis   |   Year-over-Year Trend   |   Top Products","2E75B6",10,False)

# Small segment table
dm(18,2,18,9,"Sales by Segment","2E75B6",9)
for ri2,(seg,hx) in enumerate([("Consumer","1F4E79"),("Corporate","375623"),("Home Office","7030A0")],19):
    dm(ri2,2,ri2,6,seg,hx,9,False)
    dm(ri2,7,ri2,9,"${:,.0f}".format(seg_s.get(seg,0)),"EBF3FB",9,False,color="222222")

# Small year table
dm(18,11,18,19,"Year-wise Performance","2E75B6",9)
dm(19,11,19,13,"Year","1F4E79",9)
dm(19,14,19,16,"Sales","1F4E79",9)
dm(19,17,19,19,"Profit","1F4E79",9)
for ri2,(yr,hx) in enumerate([(2020,"4472C4"),(2021,"ED7D31"),(2022,"A9D18E")],20):
    dm(ri2,11,ri2,13,str(yr),hx,9,False)
    dm(ri2,14,ri2,16,"${:,.0f}".format(yr_s[yr]),"EBF3FB",9,False,color="222222")
    dm(ri2,17,ri2,19,"${:,.0f}".format(yr_p[yr]),"EBF3FB",9,False,color="222222")

# Charts on dashboard
def mk_cat():
    c=BarChart(); c.type="col"; c.title="Sales by Category & Region"
    c.y_axis.title="Sales ($)"; c.style=10; c.width=16; c.height=12
    for idx,cat in enumerate(CATS):
        dr=Reference(ws_c,min_col=3,max_col=6,min_row=3+idx*2,max_row=3+idx*2)
        c.add_data(dr); c.series[-1].title=SeriesLabel(v=cat)
    c.set_categories(Reference(ws_c,min_col=3,max_col=6,min_row=2)); return c

def mk_yoy():
    c=LineChart(); c.title="Monthly Sales: 3 Years"
    c.y_axis.title="Sales ($)"; c.style=10; c.width=16; c.height=12
    clrs=["4472C4","ED7D31","A9D18E"]
    for ci2,(yr,clr) in enumerate(zip([2020,2021,2022],clrs),2):
        c.add_data(Reference(ws_y,min_col=ci2,max_col=ci2,min_row=2,max_row=14),titles_from_data=True)
        c.series[-1].graphicalProperties.line.solidFill=clr
        c.series[-1].graphicalProperties.line.width=25000
    c.set_categories(Reference(ws_y,min_col=1,max_col=1,min_row=3,max_row=14)); return c

def mk_top():
    c=BarChart(); c.type="bar"; c.title="Top 10 Products by Profit"
    c.style=10; c.width=16; c.height=12
    c.add_data(Reference(ws_t,min_col=3,max_col=3,min_row=2,max_row=12),titles_from_data=True)
    c.set_categories(Reference(ws_t,min_col=2,max_col=2,min_row=3,max_row=12)); return c

ws_db.add_chart(mk_cat(), "B24")
ws_db.add_chart(mk_yoy(), "S24")
ws_db.add_chart(mk_top(), "B44")

dm(68,2,68,35,"BEE Lab Project | Group 4 | B.Tech First Year | Data: retail_sales_final.csv (300 orders, 2020-2022)","1F4E79",9,False)

OUT = r"e:\BEE\Section2_Excel\Store_Sales_Dashboard_FINAL.xlsx"
wb.save(OUT)
print(f"FINAL Dashboard saved -> {OUT}")
print(f"Sheets: DASHBOARD | SalesData | PT_Category | PT_YearlyComparison | PT_TopProducts | PT_SubCategory")
print(f"Sales=${total_s:,.2f} | Profit=${total_p:,.2f} | Margin={margin:.1%} | Orders={total_o}")
print(f"2020=${s2020:,.2f} | 2021=${s2021:,.2f}(+{yoy2021:.1f}%) | 2022=${s2022:,.2f}({yoy2022:+.1f}%)")
